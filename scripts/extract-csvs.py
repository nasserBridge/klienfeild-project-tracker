"""Extract each sheet of the master tracker xlsx to a CSV in /sample-data/.

This is a dev-only script used once to seed sample CSVs that match the spec's
required filenames.
"""
import csv
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

SRC = Path(__file__).resolve().parents[1] / "sample-data" / "26003153.001A_OCSAN Tunnels_Project Tracker.xlsx"
SRC2 = Path(__file__).resolve().parents[1] / "sample-data" / "data (9).xlsx"
OUT = Path(__file__).resolve().parents[1] / "sample-data"

# Map sheet name -> output CSV filename (matches the spec)
SHEET_MAP = {
    "All-Data Export": "pmweb_all_data_export.csv",
    "Proj Trans Detail": "kfasts_proj_trans_detail.csv",
    "Invoice Summary": "invoice_summary.csv",
    "Task Summary": "task_summary.csv",
    "Task Budget": "task_budget.csv",
    "ETC": "etc.csv",
    "Invoice Log": "invoice_log.csv",
    "Change Log": "change_log.csv",
    "Sub Management": "sub_management.csv",
    "Staff ": "staff.csv",  # note trailing space in source
    "Notes": "notes.csv",
    "Check Detail": "check_detail.csv",
    "Tables": "tables.csv",
    "JTD Revenue by FY Periods": "jtd_revenue_by_period.csv",
    "Hours by Staff": "hours_by_staff.csv",
}


def write_csv(ws, dest):
    with open(dest, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            # Trim fully-empty trailing cells so files aren't ridiculously wide
            cells = list(row)
            while cells and (cells[-1] is None or cells[-1] == ""):
                cells.pop()
            if not cells:
                w.writerow([])
                continue
            w.writerow(["" if c is None else c for c in cells])


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    print(f"Sheets in {SRC.name}:")
    for sn in wb.sheetnames:
        print(f"  - {sn!r}")

    written = []
    for sheet, fname in SHEET_MAP.items():
        if sheet not in wb.sheetnames:
            print(f"!! sheet not found: {sheet!r}")
            continue
        ws = wb[sheet]
        dest = OUT / fname
        write_csv(ws, dest)
        written.append((fname, ws.max_row, ws.max_column))

    print("\nWrote:")
    for fname, r, c in written:
        print(f"  {fname}: {r} rows x {c} cols")


if __name__ == "__main__":
    main()
